from __future__ import annotations

from pathlib import Path

from models import ExtractionResult


def extract_xlsx(path: Path, max_chars: int = 8000) -> ExtractionResult:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        lines: list[str] = []

        for sheet in wb.worksheets:
            sheet_lines: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    sheet_lines.append(" | ".join(cells))
            if sheet_lines:
                lines.append(f"[시트: {sheet.title}]")
                lines.extend(sheet_lines)

        wb.close()
        combined = "\n".join(lines).strip()

        if not combined:
            return ExtractionResult(
                file_type="xlsx",
                extraction_status="empty_text",
                notes=["No text extracted from XLSX."],
            )

        return ExtractionResult(
            file_type="xlsx",
            extraction_status="success",
            extracted_text=combined,
            text_excerpt=combined[:max_chars],
            page_count=len(wb.worksheets),
            notes=[f"XLSX text extracted. {len(wb.worksheets)} sheet(s)."],
        )

    except ModuleNotFoundError:
        return ExtractionResult(
            file_type="xlsx",
            extraction_status="missing_dependency:openpyxl",
            notes=["openpyxl is not installed. Run: pip install openpyxl"],
        )
    except Exception as exc:
        return ExtractionResult(
            file_type="xlsx",
            extraction_status=f"error:{exc}",
            notes=["XLSX extraction failed."],
        )
