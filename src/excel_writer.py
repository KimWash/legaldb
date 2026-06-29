from __future__ import annotations

from pathlib import Path
from datetime import datetime

from models import AnalysisRecord


HEADERS = [
    "seq", "relative_path", "original_folder", "original_file_name", "original_full_path",
    "file_extension", "file_size", "last_modified_time", "file_type", "extraction_status",
    "ocr_used", "extracted_doc_type", "extracted_summary", "document_abstract", "extracted_document_title", "extracted_case_name", "extracted_institution",
    "extracted_date", "extracted_keyword", "suggested_file_name", "suggested_full_path", "manually_edited",
    "reason", "confidence", "needs_manual_review", "approval", "reviewer_comment",
    "rename_status", "rollback_name",
    "sharepoint_item_id", "sharepoint_web_url",
    # Legal metadata (Excel only)
    "case_name_normalized", "case_alias", "case_type", "dispute_type",
    "document_category", "document_type_normalized", "procedure_stage",
    "document_purpose", "legal_issue_primary", "legal_issue_secondary",
    "issue_tags", "claim_type", "party_our_side", "party_counterparty",
    "party_role", "law_firm_name_normalized", "institution_role",
    "country_region", "amount_mentioned", "claim_amount", "currency",
    "amount_context", "event_date", "date_type", "next_action_date",
    "timeline_summary", "lawyer_summary", "search_summary",
    "recommended_use", "review_priority", "review_priority_reason",
    "metadata_limitations", "needs_legal_review",
]


def write_review_workbook(records: list[AnalysisRecord], output_dir: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet import _writer as worksheet_writer

    def safe_cleanup(self) -> None:
        import os

        if getattr(self, "out", None) is not None:
            worksheet_writer.ALL_TEMP_FILES = [path for path in worksheet_writer.ALL_TEMP_FILES if path != self.out]
            try:
                os.remove(self.out)
            except PermissionError:
                pass

    worksheet_writer.WorksheetWriter.cleanup = safe_cleanup

    output_path = output_dir / f"rename_review_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "rename_review"
    sheet.append(HEADERS)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    warning_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    conflict_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    manual_edit_fill = PatternFill(fill_type="solid", fgColor="D6E4F0")
    manual_edit_font = Font(bold=True, color="1A5276")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    from openpyxl.styles import Alignment
    abstract_col = HEADERS.index("document_abstract") + 1
    timeline_col = HEADERS.index("timeline_summary") + 1
    lawyer_col = HEADERS.index("lawyer_summary") + 1
    search_col = HEADERS.index("search_summary") + 1
    _wrap_cols = (abstract_col, timeline_col, lawyer_col, search_col)

    for record in records:
        row = record.to_excel_row()
        sheet.append([row.get(header, "") for header in HEADERS])
        row_index = sheet.max_row
        for col in _wrap_cols:
            sheet.cell(row=row_index, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        if bool(row["needs_manual_review"]):
            for cell in sheet[row_index]:
                cell.fill = warning_fill
        if str(row["rename_status"]) == "duplicate_conflict":
            for cell in sheet[row_index]:
                cell.fill = conflict_fill
        if row.get("manually_edited") == "✓":
            me_col = HEADERS.index("manually_edited") + 1
            sf_col = HEADERS.index("suggested_file_name") + 1
            sheet.cell(row=row_index, column=me_col).fill = manual_edit_fill
            sheet.cell(row=row_index, column=me_col).font = manual_edit_font
            sheet.cell(row=row_index, column=sf_col).fill = manual_edit_fill
            sheet.cell(row=row_index, column=sf_col).font = manual_edit_font

    validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    sheet.add_data_validation(validation)
    approval_col = HEADERS.index("approval") + 1
    validation.add(f"{sheet.cell(row=2, column=approval_col).coordinate}:{sheet.cell(row=max(sheet.max_row, 2), column=approval_col).coordinate}")

    col_widths = {
        "B": 45, "C": 35, "D": 45, "E": 70, "L": 20, "M": 28, "N": 60,
        "O": 20, "R": 45, "S": 70, "T": 60, "U": 70, "V": 10, "Y": 22,
        # Legal metadata columns (AF=32 … BL=64)
        "AF": 30,  # case_name_normalized
        "AM": 40,  # document_purpose
        "AN": 35,  # legal_issue_primary
        "AO": 35,  # legal_issue_secondary
        "AP": 40,  # issue_tags
        "AR": 25,  # party_our_side
        "AS": 25,  # party_counterparty
        "AU": 25,  # law_firm_name_normalized
        "BE": 55,  # timeline_summary
        "BF": 55,  # lawyer_summary
        "BG": 55,  # search_summary
        "BJ": 35,  # review_priority_reason
        "BK": 40,  # metadata_limitations
    }
    for column, width in col_widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    workbook.save(output_path)
    return output_path


def _read_via_win32com(review_file: Path) -> list[dict]:
    """MIP 암호화 등 openpyxl이 열지 못하는 파일을 Excel COM으로 직접 읽는다.

    파일로 재저장하지 않고 COM UsedRange.Value에서 데이터를 직접 추출하므로
    MIP 정책이 임시 파일에 재적용되는 문제를 피할 수 있다.
    """
    import win32com.client  # type: ignore

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(review_file.resolve()))
        ws = wb.Worksheets(1)
        used = ws.UsedRange
        data = used.Value  # tuple of tuples (row, col)
        wb.Close(False)
    finally:
        try:
            excel.Quit()
        except Exception:
            pass

    if not data:
        return []

    # data[0] = 헤더 행
    headers = [str(cell) if cell is not None else "" for cell in data[0]]
    rows: list[dict] = []
    for row in data[1:]:
        rows.append({headers[i]: (v if v is not None else "") for i, v in enumerate(row)})
    return rows


def read_review_rows(review_file: Path) -> list[dict]:
    from openpyxl import load_workbook
    import zipfile

    # 파일 헤더로 포맷 확인
    with open(review_file, "rb") as f:
        header = f.read(4)

    is_zip = header[:2] == b"PK"

    if is_zip:
        workbook = load_workbook(review_file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        return [{headers[i]: v for i, v in enumerate(row)} for row in sheet.iter_rows(min_row=2, values_only=True)]

    # ZIP이 아니면 MIP 암호화 또는 구형 포맷 → win32com으로 시도
    print("[excel] MIP 암호화 감지 → Excel COM으로 읽기 시도...")
    try:
        return _read_via_win32com(review_file)
    except ImportError:
        raise RuntimeError(
            "pywin32가 설치되어 있지 않습니다. 'pip install pywin32' 후 다시 실행하세요.\n"
            "또는 Excel에서 MIP 레이블을 제거한 뒤 .xlsx로 저장하세요."
        )
    except Exception as exc:
        raise RuntimeError(f"Excel COM으로 파일 읽기 실패: {exc}")
