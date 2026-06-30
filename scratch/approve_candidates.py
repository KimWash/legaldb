import sys
from pathlib import Path
from openpyxl import load_workbook

excel_file = Path("d:/Legal_DB_Rename_Project_MSsharepoint/rename_review_20260630_160120.xlsx")
if not excel_file.exists():
    print(f"Error: {excel_file} not found")
    sys.exit(1)

print(f"Loading {excel_file.name}...")
wb = load_workbook(excel_file)
ws = wb.active

# Find header indexes (1-indexed for openpyxl columns)
headers = [cell.value for cell in ws[1]]
print("Headers:", headers)

idx_orig = headers.index("original_file_name") + 1
idx_sugg = headers.index("suggested_file_name") + 1
idx_status = headers.index("rename_status") + 1
idx_manual = headers.index("needs_manual_review") + 1
idx_appr = headers.index("approval") + 1
idx_sp_id = headers.index("sharepoint_item_id") + 1

print(f"Column indexes: orig={idx_orig}, sugg={idx_sugg}, status={idx_status}, manual={idx_manual}, appr={idx_appr}")

approved_count = 0
for row in range(2, ws.max_row + 1):
    orig = ws.cell(row=row, column=idx_orig).value
    sugg = ws.cell(row=row, column=idx_sugg).value
    status = ws.cell(row=row, column=idx_status).value
    manual = ws.cell(row=row, column=idx_manual).value
    sp_id = ws.cell(row=row, column=idx_sp_id).value
    
    # Check if a rename is proposed, it is not a duplicate conflict, and has a sharepoint item id
    if orig != sugg and status != "duplicate_conflict" and sp_id:
        ws.cell(row=row, column=idx_appr).value = "Y"
        approved_count += 1

print(f"Saving workbook. Total approved rows: {approved_count}")
wb.save(excel_file)
print("Workbook saved successfully!")
